import openpyxl
from datetime import datetime
import os

class ExcelHandler:
    def __init__(self, filename='crane_operations.xlsx'):
        self.filename = filename
        self.create_workbook_if_not_exists()

    def create_workbook_if_not_exists(self):
        if not os.path.exists(self.filename):
            workbook = openpyxl.Workbook()
            
            # Create sheets
            sheets = ['Crane Data', 'Barge Data', 'Generator Data', 'Ship Data']
            for sheet_name in sheets:
                if sheet_name not in workbook.sheetnames:
                    workbook.create_sheet(sheet_name)
            
            # Remove default sheet
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
            
            # Add headers to each sheet
            self.add_headers(workbook)
            
            workbook.save(self.filename)

    def add_headers(self, workbook):
        # Crane Data Sheet Headers
        crane_headers = [
            'Date', 'Shift', 'Crane Number', 'Operator', 
            'Start Time', 'Stop Time', 'Active Duration', 'Idle Reason'
        ]
        workbook['Crane Data'].append(crane_headers)

        # Barge Data Sheet Headers
        barge_headers = [
            'Date', 'Barge Name/ID', 'Start Time', 'Stop Time', 
            'Tons Loaded'
        ]
        workbook['Barge Data'].append(barge_headers)

        # Generator Data Sheet Headers
        generator_headers = [
            'Date', 'Generator ID', 'Start Time', 'Stop Time', 
            'Active Duration'
        ]
        workbook['Generator Data'].append(generator_headers)

        # Ship Data Sheet Headers
        ship_headers = [
            'Date', 'Ship Name', 'Start Time', 'Finished Time', 
            'Quantity', 'Number of Hatches'
        ]
        workbook['Ship Data'].append(ship_headers)

    def log_crane_data(self, crane_number, operator, start_time, stop_time, idle_reason):
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook['Crane Data']
        
        # Determine shift
        def determine_shift(time):
            hour = time.hour
            if 7 <= hour < 14:
                return 'Shift 1'
            elif 14 <= hour < 22:
                return 'Shift 2'
            else:
                return 'Shift 3'

        row = [
            datetime.now().date(),
            determine_shift(start_time),
            crane_number,
            operator,
            start_time.strftime('%H:%M:%S'),
            stop_time.strftime('%H:%M:%S'),
            str(stop_time - start_time),
            idle_reason
        ]
        sheet.append(row)
        workbook.save(self.filename)

    def log_barge_data(self, barge_name, start_time, stop_time, tons_loaded):
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook['Barge Data']
        
        row = [
            datetime.now().date(),
            barge_name,
            start_time.strftime('%H:%M:%S'),
            stop_time.strftime('%H:%M:%S'),
            tons_loaded
        ]
        sheet.append(row)
        workbook.save(self.filename)

    def log_generator_data(self, generator_id, start_time, stop_time):
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook['Generator Data']
        
        row = [
            datetime.now().date(),
            generator_id,
            start_time.strftime('%H:%M:%S'),
            stop_time.strftime('%H:%M:%S'),
            str(stop_time - start_time)
        ]
        sheet.append(row)
        workbook.save(self.filename)

    def log_ship_data(self, ship_name, start_time=None, finished_time=None, quantity=None, hatches=None):
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook['Ship Data']
        
        row = [
            datetime.now().date(),
            ship_name,
            start_time.strftime('%H:%M:%S') if start_time else None,
            finished_time.strftime('%H:%M:%S') if finished_time else None,
            quantity,
            hatches
        ]
        sheet.append(row)
        workbook.save(self.filename)

__all__ = ['ExcelHandler']